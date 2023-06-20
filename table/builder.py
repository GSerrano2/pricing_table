try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except:
    pass

def buildExcelTable(data):
    wb = Workbook()
    sheet = wb.active

    boldFont = Font(bold=True)

    centered = Alignment(horizontal="center", vertical="center")

    orangeFill = PatternFill(start_color="e6a02e", fill_type="solid")
    whiteFill = PatternFill(start_color="FFFFFF", fill_type="solid")

    whiteFont = Font(color="FFFFFF", bold=True)

    orangeBorder = Border(left=Side(border_style='thin', color="e6a02e"), 
                       right=Side(color="e6a02e"), 
                       top=Side(color="e6a02e"), 
                       bottom=Side(border_style='thin', color="e6a02e"))
    
    orangeBorder2 = Border(left=Side(color="e6a02e"), 
                       right=Side(border_style='thin', color="e6a02e"), 
                       top=Side(color="e6a02e"), 
                       bottom=Side(border_style='thin', color="e6a02e"))
    
    orangeBorder3 = Border(left=Side(color="e6a02e"), 
                    right=Side(color="e6a02e"), 
                    top=Side(color="e6a02e"), 
                    bottom=Side(border_style='thin', color="e6a02e"))

    # Crea la tabla de color naranja
    table_data = [
        ['Description', 'Service', 'Monthly', 'First 12 Months Total', 'Currency', 'Quantity', 'Observations'],
    ]

    rowAttr = []
    for col in range(len(data[0])):
        for row in range(len(data)):
            if row == len(data) - 1:
                lastRow = ""
                for d in data[row][col].values():
                    lastRow += str(d) + " \r\n"
                append = str(lastRow)
            else:
                append = str(data[row][col])
            rowAttr.append(append)    
        table_data.append(rowAttr)
        rowAttr = []

    table_data.append(["Total", " ", totalMontlyCosts(data), totalAnualCosts(data), "USD", "-", "-"])

    # Agrega los datos a la hoja de cálculo
    for row_data in table_data:
        sheet.append(row_data)

    # Aplica el formato a la primera fila (cabecera)
    w = 1
    header_row = sheet[1]
    for cell in header_row:
        cell.fill = orangeFill
        cell.font = whiteFont
        cell.alignment = centered     

        column = get_column_letter(w)
        sheet.column_dimensions[column].width = 20
        w += 1   

    for cell in sheet[len(table_data)]:
        cell.font = boldFont

    # Aplica el formato a las filas restantes
    for row in sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            cell.fill = whiteFill
            cell.alignment = centered
            cell.border = orangeBorder

    # Aplica el formato de borde a cada columna
    h = 1
    for row in sheet.iter_rows():
        sheet.row_dimensions[h].height = 50
        h += 1
        for i, cell in enumerate(row):
            if i == 0:
                cell.border = orangeBorder
            elif i == len(row) -1:
                cell.border = orangeBorder2
            else:
                cell.border = orangeBorder3
        
    # Guarda el libro de Excel
    wb.save('pricing_table.xlsx')
    return

def totalMontlyCosts(data):
    total = 0
    for d in data[2]:
        total += float(d.replace(",", "."))
    total = "{:,.2f}".format(total).replace('.', '_').replace(',', '.').replace('_', ',')
    return total

def totalAnualCosts(data):
    total = 0
    for d in data[3]:
        total += float(d.replace(",", "."))
    total = "{:,.2f}".format(total).replace('.', '_').replace(',', '.').replace('_', ',')
    return total